#!/usr/bin/env python
# coding: utf-8

# In[219]:


#pip install requests
#pip install html5lib
#!pip install openpyxl
#!pip install xlrd


# In[224]:


#pip install spacy
#!pip install textstat


# In[92]:

#Importing required libraries

import requests
import bs4
from bs4 import BeautifulSoup
import html
import pandas as pd
import numpy as np
import xlrd
from openpyxl import load_workbook
import nltk 
from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize
import spacy
from textstat.textstat import textstatistics
from nltk.corpus import stopwords
import string
import re

import pandas as pd


# In[110]:


# to load and extract article into text file

def extract_article_text(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find and extract the article title
    article_title = soup.find('title').get_text()

    # Find and extract the article text
    article_text = ''
    article_content = soup.find('article')
    if article_content:
        paragraphs = article_content.find_all('p')
        article_text = ' '.join([p.get_text() for p in paragraphs])

    return article_title, article_text


# Load the Excel file
workbook = load_workbook('Input.xlsx')
sheet = workbook.active
article_text = list()
# Iterate over the rows in the Excel file
for row in sheet.iter_rows(min_row=2, values_only=True):
    url_id, url = row

    # Extract the article text from the URL
    article_title, article_text = extract_article_text(url)
    tokens = word_tokenize(article_text)

print("Extraction completed successfully.")


# In[93]:


negative = open("./MasterDictionary/negative-words.txt","r")
positive = open("./MasterDictionary/positive-words.txt","r")


# In[111]:


nltk.download('punkt')
nltk.download('stopwords')


# In[95]:


with open('./MasterDictionary/negative-words.txt',"r",encoding="ISO-8859-1") as f:
    neg = f.readlines()
with open('./MasterDictionary/positive-words.txt',"r",encoding="ISO-8859-1") as f:
    pos = f.readlines()
p = len(pos)
n = len(neg)


# In[96]:


#combining all stop files into single string(data2)

data = data2 = ""
 
# Reading data from file1
with open('./StopWords/StopWords_Auditor.txt',"r",encoding="ISO-8859-1") as fp:
    data = fp.read()

# Reading data from file2
with open('./StopWords/StopWords_Currencies.txt',"r",encoding="ISO-8859-1") as fp:
    data2 = fp.read()
data +=data2
with open('./StopWords/StopWords_DatesandNumbers.txt',"r",encoding="ISO-8859-1") as fp:
    data2 = fp.read()
data +=data2
with open('./StopWords/StopWords_Generic.txt',"r",encoding="ISO-8859-1") as fp:
    data2 = fp.read()
data +=data2

with open('./StopWords/StopWords_GenericLong.txt',"r",encoding="ISO-8859-1") as fp:
    data2 = fp.read()
data +=data2

with open('./StopWords/StopWords_Geographic.txt',"r",encoding="ISO-8859-1") as fp:
    data2 = fp.read()
data +=data2

with open('./StopWords/StopWords_Names.txt',"r",encoding="ISO-8859-1") as fp:
    data2 = fp.read()
data +=data2


# In[97]:


i =0
#removing \n 
for i in range(0,p):
    pos[i] = pos[i].strip()
print(len(pos))

#removing provided stop words from given postive dictionary
pos = [word for word in pos if word.lower() not in data]    
print(len(pos))
for i in range(0,n):
    neg[i] = neg[i].strip()
print(len(neg))

#removing provided stop words from given postive dictionary
neg = [word for word in neg if word.lower() not in data]   
print(len(neg))     


# In[99]:


text_len=[]
neg_score=[]
pos_score=[]
polar_score = []
sub_score=[]
avg_sen_len=[]
compl_text=[]
fog_index=[]
avg_word_sen = []
word_clean = []
dict_syll={}
complex_word=[]
average_word_len=[]


for i in range(37,151):
    neg_sum=0
    pos_sum=0
    com_word=0
    with open(f"{i}.0.txt",'r') as f:
        f_text = f.read()
    sen_tok = sent_tokenize(f_text)
    tokens = word_tokenize(f_text)
    len_sen = len(sen_tok)
    len_word = len(tokens) 
    #removing stopwords
    tokens = [word for word in tokens if word.lower() not in data]
    
    token_not_punc =[word for word in tokens if word not in string.punctuation]
    from nltk.corpus import stopwords
    stopwords = set(stopwords.words('english'))
    token_not_punc = [word for word in token_not_punc if word.lower() not in stopwords]
    
    #length of text after cleaning
    text_len.append(len(token_not_punc))

    total_characters = sum(len(word) for word in tokens)
    
    
    
    for words in tokens:
        syll_count = textstatistics().syllable_count(words)
        
        if syll_count>2:
            #print(words)
            com_word = com_word+1
            
        #checking in positive words dictionaries
        if words in pos:
            pos_sum = pos_sum +1

        if words  in neg:
            # subracting with every negative terms in files
            neg_sum = neg_sum-1

        # multiplying with -1 to get positive score
    neg_score.append(-1*neg_sum)
    pos_score.append(pos_sum)
    polar_score.append((pos_score[k]-neg_score[k])/(pos_score[k]+neg_score[k]+0.000001))
    sub_score.append((pos_score[k]+neg_score[k])/(text_len[k]+0.000001))
    avg_sen_len.append(len_word/len_sen)
    #print(com_word)
    compl_text.append(com_word/len_word)
    #print(compl_text)
    fog_index.append(0.4*(com_word/len_word+ len_word/len_sen ))
    avg_word_sen.append(len_word/len_sen)
    word_clean.append(len(token_not_punc))
    average_word_len.append(total_characters/len(tokens))
    complex_word.append(com_word)
    
    #print(avg_word_sen[k])
    #print(word_clean)
#print(average_word_len)
   


# In[100]:


import nltk
from nltk.corpus import cmudict
import re

nltk.download('cmudict')

# can be used to count syllabus but it will take alot of computational time.

def count_syllables(word):
    pronunciation_list = cmudict.dict().get(word.lower())
    if pronunciation_list:
        pronunciation = pronunciation_list[0]
        syllable_count = 0
        for phoneme in pronunciation:
            if re.match(r'\d', phoneme[-1]):
                syllable_count += 1
        if word.endswith(('es', 'ed')):
            syllable_count -= 1
        return syllable_count
    else:
        return 0


# In[101]:

#for storing syllabus per word in text files
array_count_syll=[]
dict_syll =[]
count =0
for i in range(37,151):
    count = 0
    with open(f"{i}.0.txt",'r') as f:
        f_text = f.read()
    sen_tok = sent_tokenize(f_text)
    tokens = word_tokenize(f_text)
    for words in tokens:
        # count syllabus of every word and taking average
        count = count + textstatistics().syllable_count(words)
        
    dict_syll.append(count/len(tokens))
    
    
    
    


# In[103]:


#Counting given pronouns from text files

def count_personal_pronouns(text):
    pronouns = r'\b(I|we|my|ours|us)\b'
    excluded_words = r'\bUS\b'

    # Find matches of personal pronouns
    matches = re.findall(pronouns, text, flags=re.IGNORECASE)

    # Exclude matches that are the country name "US"
    filtered_matches = [match for match in matches if not re.match(excluded_words, match, flags=re.IGNORECASE)]

    return len(filtered_matches)

pers_pro = []
for i in range(37,151):
    with open(f"{i}.0.txt",'r') as f:
        f_text = f.read()

    pers_pro.append(count_personal_pronouns(f_text))


    


# In[105]:


#converting all arrays to dataframes to replace it with empty coloumns in Output Data Structure.xlsx

output = pd.read_excel("Output Data Structure.xlsx")
pos_score = pd.DataFrame(pos_score)
neg_score = pd.DataFrame(neg_score)
polar_score = pd.DataFrame(polar_score)
sub_score = pd.DataFrame(sub_score)
avg_sen_len = pd.DataFrame(avg_sen_len)
compl_text= pd.DataFrame(compl_text)
fog_index = pd.DataFrame(fog_index)
avg_word_sen = pd.DataFrame(avg_word_sen)
complex_word = pd.DataFrame(complex_word)
text_len = pd.DataFrame(text_len)
pers_pro = pd.DataFrame(pers_pro)
average_word_len = pd.DataFrame(average_word_len)
array_count_syll = pd.DataFrame(array_count_syll)
dict_syll = pd.DataFrame(dict_syll)


# In[108]:


output["POSITIVE SCORE"] = pos_score.values
output["NEGATIVE SCORE"] = neg_score.values
output["POLARITY SCORE"] = polar_score.values
output["SUBJECTIVITY SCORE"]=sub_score.values
output["AVG SENTENCE LENGTH"] = avg_sen_len.values
output["PERCENTAGE OF COMPLEX WORDS"]= compl_text.values
output["FOG INDEX"] = fog_index.values
output["AVG NUMBER OF WORDS PER SENTENCE"] = avg_word_sen.values
output["COMPLEX WORD COUNT"] = complex_word.values
output["WORD COUNT"]= text_len.values
output["PERSONAL PRONOUNS"] = pers_pro.values
output["AVG WORD LENGTH"] = average_word_len.values
output["SYLLABLE PER WORD"] = dict_syll
#output["SYLLABLE PER WORD"] = dict_syll
output.to_excel("Tast_completed.xlsx")







